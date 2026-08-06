"""
语法测试：确认 officecli 列删除 (col[]) 和合并单元格 (merge prop) 的正确写法。
仅生成 regression_ 前缀测试产物。
"""
import subprocess, shutil, os
from openpyxl import load_workbook

REF_DIR = r"c:\Users\59620\.claude\skills\tendo-brand\references"
TEMPLATE = REF_DIR + r"\TendoCN - Cooley LLP - Cooley Shanghai Meeting Room Retrofit - Weekly Progress Report (项目周报) .xlsx"
OUTPUT = REF_DIR + r"\regression_syntax_test.xlsx"

def oc(args, verbose=True):
    cmd = ["officecli"] + args
    r = subprocess.run(cmd, capture_output=True, text=True)
    status = "OK" if r.returncode == 0 else f"FAIL({r.returncode})"
    if verbose:
        print(f"  [{status}] officecli " + " ".join(str(a) for a in args))
        if r.stdout.strip(): print(f"         out: {r.stdout.strip()[:200]}")
        if r.returncode != 0 and r.stderr.strip(): print(f"         err: {r.stderr.strip()[:200]}")
    return r.returncode == 0, r.stdout, r.stderr

# 先关闭可能存在的 resident
oc(["close", OUTPUT], verbose=False)
shutil.copy(TEMPLATE, OUTPUT)
print(f"复制模板 → {OUTPUT}\n")

# ========== 测试 1: col[] 删除 ==========
print("=" * 60)
print("测试 1: 列删除 col[] 语法 (删 Progress Report 第 9 列)")
print("=" * 60)
# 1a: col[9] (index)
print("\n[1a] col[9] (数字 index):")
oc(["remove", OUTPUT, "/Progress Report/col[9]"])
# 1b: col[I] (letter, I=第9列)
print("\n[1b] col[I] (字母):")
oc(["remove", OUTPUT, "/Progress Report/col[I]"])
# 1c: column[9]
print("\n[1c] column[9]:")
oc(["remove", OUTPUT, "/Progress Report/column[9]"])

# 检查是否真的删了 (close 后读)
oc(["close", OUTPUT], verbose=False)
wb = load_workbook(OUTPUT)
ws = wb["Progress Report"]
# 模板 row 13: C13=Demolition, D13=Demolition子列, I13=Restoration
# 删 col 9 (I) 后, 原 J 左移到 I, 所以 I13 应该变成 New Point Wiring (原 L13) 或别的
c13 = ws.cell(row=13, column=3).value  # C
i13 = ws.cell(row=13, column=9).value  # I (col 9)
print(f"\n  验证: C13={c13!r}, I13={i13!r}")
print(f"  (模板原 I13='Restoration'; 若删除成功 I13 会变成原 J13 的内容)")
wb.close()

# 重新复制模板做 merge 测试
oc(["close", OUTPUT], verbose=False)
shutil.copy(TEMPLATE, OUTPUT)

# ========== 测试 2: merge prop 合并单元格 ==========
print("\n" + "=" * 60)
print("测试 2: 合并单元格 merge prop 语法 (合并 Issue_RFA Log C14:D14)")
print("=" * 60)
# 先确认 C14:D14 在模板里已合并
wb = load_workbook(OUTPUT)
before = "C14:D14" in [str(m) for m in wb["Issue_RFA Log"].merged_cells.ranges]
wb.close()
print(f"\n  模板原始 C14:D14 合并状态: {before}")

# 用 set merge prop 合并 E14:F14 (模板里没合并的)
print("\n[2a] set E14 --prop merge=E14:F14:")
oc(["set", OUTPUT, "/Issue_RFA Log/E14", "--prop", "merge=E14:F14"])

# 用 set range merge=true
print("\n[2b] set E15:F15 --prop merge=true:")
oc(["set", OUTPUT, "/Issue_RFA Log/E15:F15", "--prop", "merge=true"])

# 用错误的 merge 子命令 (Bug 3 修复用的写法)
print("\n[2c] (Bug3 修复写法) merge 子命令:")
oc(["merge", OUTPUT, "/Issue_RFA Log/E16:F16"])

oc(["close", OUTPUT], verbose=False)
wb = load_workbook(OUTPUT)
ws = wb["Issue_RFA Log"]
merges = [str(m) for m in ws.merged_cells.ranges]
print(f"\n  验证合并结果:")
print(f"  E14:F14 merged = {'E14:F14' in merges}  (期望 True, set merge prop 生效)")
print(f"  E15:F15 merged = {'E15:F15' in merges}  (期望 True, range merge=true 生效)")
print(f"  E16:F16 merged = {'E16:F16' in merges}  (期望 False, merge 子命令无效)")
wb.close()

print("\n" + "=" * 60)
print("语法测试完成")
print("=" * 60)
