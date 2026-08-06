"""
终审模板验证脚本 — 用 openpyxl 读取真实模板值，核对修复声明。
仅读取，不修改任何文件。
"""
import sys
from openpyxl import load_workbook

REF_DIR = r"c:\Users\59620\.claude\skills\tendo-brand\references"
WORKER_LIST = REF_DIR + r"\TendoCN - Worker Name List.xlsx"
WEEKLY = REF_DIR + r"\TendoCN - Cooley LLP - Cooley Shanghai Meeting Room Retrofit - Weekly Progress Report (项目周报) .xlsx"

def banner(title):
    print("\n" + "=" * 70)
    print(title)
    print("=" * 70)

# ============ 1. Worker List 模板 ============
banner("[1] Worker List 模板 — 验证 Bug 1: row 12-15 序号 3-6, row 16-23 无序号")
wb = load_workbook(WORKER_LIST)
ws = wb.active
print(f"Sheet: {ws.title}, max_row={ws.max_row}, max_col={ws.max_column}")
print("\nColumn A (No.) row 8-25:")
for r in range(8, 26):
    val = ws.cell(row=r, column=1).value
    print(f"  A{r}: {val!r}")

# 合并单元格
print(f"\nMerged ranges (first 10): {list(ws.merged_cells.ranges)[:10]}")

# 验证 row 12-15 = 3,4,5,6
seq_12_15 = [ws.cell(row=r, column=1).value for r in range(12, 16)]
print(f"\n>>> A12:A15 = {seq_12_15}  (期望 [3,4,5,6])")
print(f">>> Bug1 PASS = {seq_12_15 == [3,4,5,6]}")

# 验证 row 16-23 无序号
seq_16_23 = [ws.cell(row=r, column=1).value for r in range(16, 24)]
print(f">>> A16:A23 = {seq_16_23}  (期望全 None)")
print(f">>> Bug1 (row16-23 empty) PASS = {all(v is None for v in seq_16_23)}")
wb.close()

# ============ 2. Weekly Report — Issue_RFA Log ============
banner("[2] Issue_RFA Log 模板 — 验证 Bug 3: I13 / I22 表头文本")
wb = load_workbook(WEEKLY)
print(f"Sheets: {wb.sheetnames}")
ws = wb["Issue_RFA Log"]
print(f"\nIssue_RFA Log: max_row={ws.max_row}, max_col={ws.max_column}")

print("\nI 列 (col 9) row 9-25:")
for r in range(9, 26):
    val = ws.cell(row=r, column=9).value
    print(f"  I{r}: {val!r}")

print("\nRow 13 全列表头 (A-J):")
for c in range(1, 11):
    from openpyxl.utils import get_column_letter
    val = ws.cell(row=13, column=c).value
    print(f"  {get_column_letter(c)}13: {val!r}")

print("\nRow 22 全列表头 (A-J) — RFA 表头默认位置:")
for c in range(1, 11):
    from openpyxl.utils import get_column_letter
    val = ws.cell(row=22, column=c).value
    print(f"  {get_column_letter(c)}22: {val!r}")

i13 = ws.cell(row=13, column=9).value
i22 = ws.cell(row=22, column=9).value
print(f"\n>>> I13 = {i13!r}  (期望 'Issue Open / Closed')")
print(f">>> I13 PASS = {i13 == 'Issue Open / Closed'}")
print(f">>> I22 = {i22!r}  (期望 'Issue Open / Closed')")
print(f">>> I22 PASS = {i22 == 'Issue Open / Closed'}")

# 验证 C:D 合并 (Issue 行 row 14-17)
print("\n合并单元格 (含 C:D 的):")
cd_merges = [str(m) for m in ws.merged_cells.ranges if 'C' in str(m) or 'D' in str(m)]
for m in cd_merges[:20]:
    print(f"  {m}")

# 检查 C14:D14, C15:D15, C16:D16, C17:D17
issue_merges = []
for r in range(14, 18):
    key = f"C{r}:D{r}"
    found = key in [str(m) for m in ws.merged_cells.ranges]
    issue_merges.append((key, found))
    print(f">>> {key} merged = {found}")
print(f">>> Issue C:D merge (14-17) PASS = {all(f for _, f in issue_merges)}")

# ============ 3. Weekly Report — Progress Report ============
banner("[3] Progress Report 模板 — 验证 Bug 2: Overall 公式行结构")
ws = wb["Progress Report"]
print(f"Progress Report: max_row={ws.max_row}, max_col={ws.max_column}")

print("\nB/X 列 row 15-25 (定位 Overall 标题行/公式行):")
from openpyxl.utils import get_column_letter
for r in range(15, 26):
    b = ws.cell(row=r, column=2).value
    c = ws.cell(row=r, column=3).value
    x = ws.cell(row=r, column=24).value  # X = col 24
    print(f"  row {r}: B={b!r} | C={c!r} | X={x!r}")

# 验证 row 22 = Overall Percentage 标题, row 23 = AVERAGE 公式
b22 = ws.cell(row=22, column=2).value
c23 = ws.cell(row=23, column=3).value
print(f"\n>>> B22 = {b22!r}  (期望 'Overall Percentage (%)')")
print(f">>> B22 PASS = {b22 == 'Overall Percentage (%)'}")
print(f">>> C23 = {c23!r}  (期望含 AVERAGE)")
print(f">>> C23 PASS = {c23 is not None and 'AVERAGE' in str(c23)}")

# 验证 row 17-20 是示例子项
print("\nB17:B20 (示例子项):")
for r in range(17, 21):
    print(f"  B{r}: {ws.cell(row=r, column=2).value!r}")

wb.close()
print("\n" + "=" * 70)
print("模板验证完成")
print("=" * 70)
