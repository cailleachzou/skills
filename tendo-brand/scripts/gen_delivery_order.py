import openpyxl
from openpyxl.styles import Alignment
import shutil
import os
import sys
from datetime import date, datetime

def parse_date(val):
    """Parse date string or return datetime as-is. Accepts YYYY-MM-DD or DD/MM/YYYY."""
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
    return datetime.today()

def gen_delivery_order(data, output_dir):
    template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'references',
                                  'TCMR2603-00005- Material Requisition -  TCSO2603-00085.xlsx')

    req_no = data["requisition_no"]
    so_no = data["sales_order_no"]
    filename = f"{req_no}- Material Requisition - {so_no}.xlsx"
    output_path = os.path.join(output_dir, filename)
    os.makedirs(output_dir, exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Material Requisition"]

    # Preserve G5 date format before writing
    g5_numfmt = ws["G5"].number_format

    # Header fields — wrap text for long content
    wrap = Alignment(wrap_text=True, vertical="center")
    ws["C4"] = data.get("deliver_to", "")
    ws["C5"] = data["company"]
    ws["C5"].alignment = wrap
    ws["C6"] = data["address"]
    ws["C6"].alignment = wrap
    ws["C7"] = data["sales_order_no"]
    ws["C8"] = data["sales_quotation_no"]
    ws["C9"] = data.get("submitted_by", "Cailleach.Zou")
    ws["G4"] = data["requisition_no"]
    ws["G5"] = parse_date(data.get("date", date.today()))
    ws["G5"].number_format = g5_numfmt
    ws["G6"] = data.get("currency", "CNY")

    # Items (start from row 12)
    for i, item in enumerate(data["items"]):
        row = 12 + i
        ws.cell(row=row, column=2, value=item["part_no"])
        desc_cell = ws.cell(row=row, column=3, value=item["description"])
        desc_cell.alignment = wrap
        ws.cell(row=row, column=5, value=item["qty"])
        ws.cell(row=row, column=6, value=item["unit"])
        ws.cell(row=row, column=7, value=item["unit_cost"])

    wb.save(output_path)
    return output_path

if __name__ == "__main__":
    test_data = {
        "requisition_no": "TCMR2607-00010",
        "company": "Cooley LLP Shanghai Representative Office",
        "address": "IFC - Tower 2 Level 35, Unit 3510, 8 Century Avenue, Pudong New Area, Shanghai, 200120",
        "sales_order_no": "TCSO2607-00110",
        "sales_quotation_no": "TCSQ2607-00184R2",
        "submitted_by": "Cailleach.Zou",
        "date": "2026-07-22",
        "currency": "CNY",
        "items": [
            {"part_no": "760191940", "description": "Faceplate 2-Port, White", "qty": 4, "unit": "个", "unit_cost": 10.00},
        ]
    }
    
    output = gen_delivery_order(test_data, r"C:\Users\59620\.claude\skills\tendo-brand\test")
    print(f"Generated: {output}")
