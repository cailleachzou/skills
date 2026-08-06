"""
回归测试 v2 (修正版): 用正确语法验证 Bug 2/3/4 核心逻辑。
- N=7 (跳过 Step 1 列删除, Overall 列 = X col 24, 与模板一致)
- M=3 (验证 Step 10 动态行号: M=3 → Overall 公式行 row 22)
- 1 issue + 1 rfa (验证 Bug 3: 用正确 merge 语法 + I 列表头文本)
- 加 officecli close (确保落盘)
关键: 此脚本用"正确语法"验证修复意图能否实现, 证明动态行号逻辑正确。
agent 文件本身的语法错误 (merge 子命令 / col[数字]) 单独在报告里指出。
"""
import subprocess, shutil
from openpyxl import load_workbook

REF_DIR = r"c:\Users\59620\.claude\skills\tendo-brand\references"
TEMPLATE = REF_DIR + r"\TendoCN - Cooley LLP - Cooley Shanghai Meeting Room Retrofit - Weekly Progress Report (项目周报) .xlsx"
OUTPUT = REF_DIR + r"\regression_weekly_test.xlsx"

def oc(args):
    cmd = ["officecli"] + args
    r = subprocess.run(cmd, capture_output=True, text=True)
    status = "OK" if r.returncode == 0 else f"FAIL({r.returncode})"
    print(f"  [{status}] " + " ".join(str(a) for a in args))
    if r.returncode != 0 and r.stderr.strip():
        print(f"         err: {r.stderr.strip()[:200]}")
    return r.returncode == 0

oc(["close", OUTPUT])
shutil.copy(TEMPLATE, OUTPUT)
print(f"复制模板 → {OUTPUT}\n")

PR = "/Progress Report"
IR = "/Issue_RFA Log"

# ========== Progress Report (N=7, M=3) ==========
print("=== Progress Report (N=7 阶段保留, M=3 子项) ===")
# Step 1: N=7 情况 C, 无需调列, overall_col = X (col 24)
# Step 2: 删 row 17-20 逆序
print("Step 2: 删示例子项行 row 17-20 (逆序)")
for r in [20, 19, 18, 17]:
    oc(["remove", OUTPUT, f"{PR}/row[{r}]"])
# Step 3: C12 标题
oc(["set", OUTPUT, f"{PR}/C12", "--prop", "value=Test Project"])
# Step 6: B16 楼层 (验证 row 16 保留)
oc(["set", OUTPUT, f"{PR}/B16", "--prop", "value=L28"])
# Step 7: 插入 3 子项行
sub_items = ["Reception", "Open Office", "Meeting Room"]
for i, s in enumerate(sub_items):
    row = 17 + i
    oc(["add", OUTPUT, PR, "--type", "row", "--index", str(row)])
    oc(["set", OUTPUT, f"{PR}/A{row}", "--prop", f"value={i+1}"])
    oc(["set", OUTPUT, f"{PR}/B{row}", "--prop", f"value={s}"])
# Step 8: 填充进度 (Reception Cable Pulling 100%)
oc(["set", OUTPUT, f"{PR}/C17", "--prop", "value=1"])
# Step 9: Overall 列 (N=7 → overall_col = X = col 24) AVERAGE 公式
for row in [17, 18, 19]:
    oc(["set", OUTPUT, f"{PR}/X{row}", "--prop", f"value==AVERAGE(C{row},F{row},I{row},L{row},O{row},R{row},U{row})"])
# Step 10: Overall 行 AVERAGE (M=3 → title=21, formula=22, range_end=21)
M = 3
overall_title_row = 18 + M   # 21
overall_formula_row = 19 + M  # 22
range_end = overall_title_row  # 21
print(f"Step 10: M=3 → overall_title_row={overall_title_row}, overall_formula_row={overall_formula_row}, range_end={range_end}")
for phase_col in ["C", "F", "I", "L", "O", "R", "U"]:
    oc(["set", OUTPUT, f"{PR}/{phase_col}{overall_formula_row}", "--prop", f"value==AVERAGE({phase_col}15:{phase_col}{range_end})"])
oc(["set", OUTPUT, f"{PR}/X{overall_formula_row}", "--prop", f"value==AVERAGE(X15:X{range_end})"])

# ========== Issue_RFA Log (1 issue, 1 rfa) ==========
print("\n=== Issue_RFA Log (1 issue, 1 rfa) ===")
# Step 1: 删 row 14 × 4
print("Step 1: 删示例 Issue 行 row[14] × 4")
for _ in range(4):
    oc(["remove", OUTPUT, f"{IR}/row[14]"])
# Step 2: 插入 1 Issue 行 (row 14) + 用正确语法合并 C:D
oc(["add", OUTPUT, IR, "--type", "row", "--index", "14"])
oc(["set", OUTPUT, f"{IR}/A14", "--prop", "value=1"])
oc(["set", OUTPUT, f"{IR}/B14", "--prop", "value=2026-07-30"])
oc(["set", OUTPUT, f"{IR}/C14", "--prop", "value=Test issue description"])
# ★ 用正确语法合并 C14:D14 (agent 文件写的是错的 merge 子命令, 这里用 set --prop merge=)
oc(["set", OUTPUT, f"{IR}/C14", "--prop", "merge=C14:D14"])
oc(["set", OUTPUT, f"{IR}/E14", "--prop", "value=Medium"])
oc(["set", OUTPUT, f"{IR}/F14", "--prop", "value=Test solution"])
oc(["set", OUTPUT, f"{IR}/G14", "--prop", "value=Cailleach"])
oc(["set", OUTPUT, f"{IR}/I14", "--prop", "value=Open"])
# Step 3: RFA title (rfa_title_row = 14 + 1 = 15)
rfa_title_row = 15
oc(["set", OUTPUT, f"{IR}/A{rfa_title_row}", "--prop", "value=RFI / RFA Log"])
oc(["set", OUTPUT, f"{IR}/A17", "--prop", "value=Project :"])
oc(["set", OUTPUT, f"{IR}/C17", "--prop", "value=Test Project"])
# Step 4: RFA 表头 (rfa_header_row = 19)
rfa_header_row = 19
oc(["set", OUTPUT, f"{IR}/A{rfa_header_row}", "--prop", "value=Item No."])
oc(["set", OUTPUT, f"{IR}/B{rfa_header_row}", "--prop", "value=Issued Date"])
oc(["set", OUTPUT, f"{IR}/C{rfa_header_row}", "--prop", "value=RFI / RFA"])
oc(["set", OUTPUT, f"{IR}/E{rfa_header_row}", "--prop", "value=Description"])
oc(["set", OUTPUT, f"{IR}/F{rfa_header_row}", "--prop", "value=Issued to"])
oc(["set", OUTPUT, f"{IR}/G{rfa_header_row}", "--prop", "value=Respond by"])
oc(["set", OUTPUT, f"{IR}/I{rfa_header_row}", "--prop", "value=Issue Open / Closed"])
oc(["set", OUTPUT, f"{IR}/J{rfa_header_row}", "--prop", "value=Remarks"])
# Step 5: 插入 1 RFA 数据行 (row 20)
oc(["add", OUTPUT, IR, "--type", "row", "--index", "20"])
oc(["set", OUTPUT, f"{IR}/A20", "--prop", "value=1"])
oc(["set", OUTPUT, f"{IR}/C20", "--prop", "value=RFI"])
oc(["set", OUTPUT, f"{IR}/I20", "--prop", "value=Open"])

# ★ 关键: close 确保落盘
print("\n★ officecli close 确保落盘")
oc(["close", OUTPUT])

# ========== 验证 ==========
print("\n" + "=" * 70)
print("验证结果 (openpyxl 读取落盘文件)")
print("=" * 70)
wb = load_workbook(OUTPUT)
ws = wb["Progress Report"]
print("\n--- Progress Report 关键行 (B/C/X 列, X=Overall) ---")
for r in [16, 17, 18, 19, 20, 21, 22, 23]:
    b = ws.cell(row=r, column=2).value
    c = ws.cell(row=r, column=3).value
    x = ws.cell(row=r, column=24).value
    print(f"  row {r}: B={b!r} | C={c!r} | X={x!r}")

c22 = ws.cell(row=22, column=3).value
x22 = ws.cell(row=22, column=24).value
b16 = ws.cell(row=16, column=2).value
b17 = ws.cell(row=17, column=2).value
print(f"\n[Bug 2] Step 10 动态行号 (M=3):")
print(f"  C22 = {c22!r}  (期望 '=AVERAGE(C15:C21)')  → {'PASS' if c22 == '=AVERAGE(C15:C21)' else 'FAIL'}")
print(f"  X22 = {x22!r}  (期望 '=AVERAGE(X15:X21)')  → {'PASS' if x22 == '=AVERAGE(X15:X21)' else 'FAIL'}")
print(f"[Bug 4] row 16 楼层行保留 (删 17-20 不动 16):")
print(f"  B16 = {b16!r}  (期望 'L28')  → {'PASS' if b16 == 'L28' else 'FAIL'}")
print(f"  B17 = {b17!r}  (期望 'Reception', 证明子项插入 row 17)  → {'PASS' if b17 == 'Reception' else 'FAIL'}")

ws = wb["Issue_RFA Log"]
print("\n--- Issue_RFA Log 关键行 (A/C/I 列) ---")
for r in [13, 14, 15, 17, 19, 20]:
    a = ws.cell(row=r, column=1).value
    c = ws.cell(row=r, column=3).value
    i = ws.cell(row=r, column=9).value
    print(f"  row {r}: A={a!r} | C={c!r} | I={i!r}")

cd14 = "C14:D14" in [str(m) for m in ws.merged_cells.ranges]
i13 = ws.cell(row=13, column=9).value
i19 = ws.cell(row=19, column=9).value
print(f"\n[Bug 3] Issue_RFA Log (用正确 merge 语法验证修复意图):")
print(f"  C14:D14 merged = {cd14}  (期望 True)  → {'PASS' if cd14 else 'FAIL'}")
print(f"  I13 (Issue 表头保留) = {i13!r}  → {'PASS' if i13 == 'Issue Open / Closed' else 'FAIL'}")
print(f"  I19 (RFA 表头, rfa_header_row=14+1+4=19) = {i19!r}  → {'PASS' if i19 == 'Issue Open / Closed' else 'FAIL'}")

wb.close()
print("\n" + "=" * 70)
print("回归测试 v2 完成")
print("=" * 70)
