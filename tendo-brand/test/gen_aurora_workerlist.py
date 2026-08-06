"""Generate TendoCN Worker Name List for Aurora Tech L28 Smart Office Retrofit."""
import openpyxl
from pathlib import Path

BASE = Path(r"c:\Users\59620\.claude\skills\tendo-brand")
TEMPLATE = BASE / "references" / "TendoCN - Worker Name List.xlsx"
OUTPUT = BASE / "test" / "TendoCN - Worker Name List - Aurora Tech L28.xlsx"

# Team members (Aurora Tech L28 Smart Office Retrofit)
TEAM = [
    {"first": "Dayne",    "last": "Chea",  "gender": "M", "designation": "Project Director",            "mobile": "+86-138-0000-0001", "email": "dayne.chea@tendo.technology",    "id": "E00123456", "nationality": "Singapore"},
    {"first": "Cailleach","last": "Zou",   "gender": "M", "designation": "Senior Project Engineer",     "mobile": "+86-138-0000-0002", "email": "cailleach.zou@tendo.technology", "id": "E00234567", "nationality": "P.R.China"},
    {"first": "ShiHao",   "last": "Liu",   "gender": "M", "designation": "Assistant Project Engineer",  "mobile": "+86-138-0000-0003", "email": "shihao.liu@tendo.technology",    "id": "E00345678", "nationality": "P.R.China"},
    {"first": "Susie",    "last": "He",    "gender": "F", "designation": "Sales Manager",               "mobile": "+86-138-0000-0004", "email": "susie.he@tendo.technology",      "id": "E00456789", "nationality": "P.R.China"},
    {"first": "Wei",      "last": "Zhang", "gender": "M", "designation": "ELV Technician",               "mobile": "+86-138-0000-0005", "email": "wei.zhang@tendo.technology",     "id": "E00567890", "nationality": "P.R.China"},
]

wb = openpyxl.load_workbook(TEMPLATE)
ws = wb.active

# --- Clear sample data + pre-filled No. in data area (rows 10..max_row, cols 1..10) ---
# Preserves formatting (column widths, borders, fonts); only values are wiped.
for r in range(10, ws.max_row + 1):
    for c in range(1, 11):
        ws.cell(row=r, column=c).value = None

# --- Write 5 team members starting at row 10 ---
for i, m in enumerate(TEAM):
    r = 10 + i
    ws.cell(row=r, column=1, value=i + 1)              # No.
    ws.cell(row=r, column=2, value=m["first"])         # First Name
    ws.cell(row=r, column=3, value=m["last"])          # Last Name
    ws.cell(row=r, column=4, value=m["gender"])        # Gender
    ws.cell(row=r, column=5, value=m["designation"])   # Designation
    ws.cell(row=r, column=6, value=m["mobile"])        # Mobile No.
    ws.cell(row=r, column=7, value=m["email"])         # Email
    ws.cell(row=r, column=8, value=m["id"])            # National ID/Passport No.
    ws.cell(row=r, column=9, value=m["nationality"])   # Nationality
    ws.cell(row=r, column=10, value=None)              # Remarks (optional, blank)

wb.save(OUTPUT)
print("Saved:", OUTPUT)
