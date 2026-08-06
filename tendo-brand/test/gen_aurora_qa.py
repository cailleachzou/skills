"""Generate Aurora Tech L28 Smart Office Retrofit Q&A sheet from template."""
import shutil
from pathlib import Path
import openpyxl

BASE = Path(r"c:\Users\59620\.claude\skills\tendo-brand")
TEMPLATE = BASE / "references" / "TendoCN - Q&A for (Client Name) (Project Name) - (Date).xlsx"
OUT_DIR = BASE / "test"
OUT_NAME = "TendoCN - Q&A for Aurora Tech L28 Smart Office Retrofit - 30th Jul'26.xlsx"
OUTPUT = OUT_DIR / OUT_NAME

PROJECT_NAME = "Aurora Tech Shanghai L28 Smart Office Retrofit"
SURVEY_DATE = "Survey Date: 30th Jul'26"

# Step 1: 复制模板（绝对禁止从零创建 Workbook）
shutil.copy(TEMPLATE, OUTPUT)

# Step 2: 打开复制文件（保留所有格式）
wb = openpyxl.load_workbook(OUTPUT)
ws = wb.active

# Step 3: 只写值，不动任何样式属性
ws['A3'] = PROJECT_NAME
ws['A4'] = SURVEY_DATE

wb.save(OUTPUT)
print(f"Generated: {OUTPUT}")
