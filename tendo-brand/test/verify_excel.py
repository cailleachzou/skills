import openpyxl
wb = openpyxl.load_workbook(r'c:\Users\59620\.claude\skills\tendo-brand\test\TCMR2607-00010- Material Requisition - TCSO2607-00110.xlsx')
ws = wb['Material Requisition']

print('Header:')
print('C4 Deliver To:', repr(ws['C4'].value))
print('C5 Company:', repr(ws['C5'].value))
print('C6 Address:', repr(ws['C6'].value))
print('C7 Sales Order:', repr(ws['C7'].value))
print('C8 Quotation:', repr(ws['C8'].value))
print('C9 Submitted By:', repr(ws['C9'].value))
print('G4 Req No:', repr(ws['G4'].value))
print('G5 Date:', repr(ws['G5'].value))
print('G6 Currency:', repr(ws['G6'].value))

print('\nItem Row 12:')
print('B12 Part No:', repr(ws['B12'].value))
print('C12 Description:', repr(ws['C12'].value))
print('E12 Qty:', repr(ws['E12'].value))
print('F12 Unit:', repr(ws['F12'].value))
print('G12 Unit Cost:', repr(ws['G12'].value))
print('H12 Total (formula):', repr(ws['H12'].value))

print('\nFormulas preserved:')
for cell in ['A13', 'H13', 'G7', 'G8', 'G9']:
    v = ws[cell].value
    print(f'{cell}: {repr(v)}')
