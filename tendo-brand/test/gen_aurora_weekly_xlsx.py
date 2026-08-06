"""Generate Aurora Tech L28 weekly progress report (xlsx) via openpyxl.

Implements the logic described in agents/weekly-report*.md (Progress Report,
Site Photo, Issue_RFA Log sheets) using openpyxl instead of officecli.
"""
import os
import shutil
from datetime import datetime

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SKILL_DIR = r'c:\Users\59620\.claude\skills\tendo-brand'
TEMPLATE = os.path.join(
    SKILL_DIR, 'references',
    'TendoCN - Cooley LLP - Cooley Shanghai Meeting Room Retrofit - Weekly Progress Report (项目周报) .xlsx')
OUTPUT = os.path.join(
    SKILL_DIR, 'test',
    'TendoCN - Aurora Tech Shanghai - L28 Smart Office Retrofit Weekly Progress Report (项目周报) 2026-07-30.xlsx')

# ---- Project data (Aurora Tech L28 Smart Office Retrofit) ----
PROJECT_TITLE = 'Aurora Tech Shanghai - L28 Smart Office Retrofit'
PROJECT_FLOOR = 'L28'
REPORT_DATE = '2026-07-30'
PHASES = ['Cable Pulling', 'Termination', 'Faceplate Installation', 'Testing']
SUB_ITEMS = ['Reception', 'Open Office', 'Meeting Room']

PROGRESS = {
    'Reception': {
        'Cable Pulling':         {'pct': 100, 'status': 'Completed',    'till': '2026-07-25', 'target': '2026-07-25'},
        'Termination':           {'pct': 80,  'status': 'In Progress',  'till': '2026-07-28', 'target': '2026-07-30'},
        'Faceplate Installation':{'pct': 50,  'status': 'In Progress',  'till': '2026-07-29', 'target': '2026-07-31'},
        'Testing':               {'pct': 0,   'status': 'Not Started',  'till': '',           'target': '2026-08-02'},
    },
    'Open Office': {
        'Cable Pulling':         {'pct': 80,  'status': 'In Progress',  'till': '2026-07-28', 'target': '2026-07-30'},
        'Termination':           {'pct': 30,  'status': 'In Progress',  'till': '2026-07-29', 'target': '2026-08-01'},
        'Faceplate Installation':{'pct': 0,   'status': 'Not Started',  'till': '',           'target': '2026-08-03'},
        'Testing':               {'pct': 0,   'status': 'Not Started',  'till': '',           'target': '2026-08-05'},
    },
    'Meeting Room': {
        'Cable Pulling':         {'pct': 20,  'status': 'Delay',        'till': '2026-07-29', 'target': '2026-07-28'},
        'Termination':           {'pct': 0,   'status': 'Not Started',  'till': '',           'target': '2026-08-02'},
        'Faceplate Installation':{'pct': 0,   'status': 'Not Started',  'till': '',           'target': '2026-08-04'},
        'Testing':               {'pct': 0,   'status': 'Not Started',  'till': '',           'target': '2026-08-06'},
    },
}

STATUS_FONT_COLOR = {
    'In Progress': 'FF00B050',
    'Delay':       'FFFF0000',
    'Not Started': 'FFFFC000',
    'Completed':   'FF000000',
}
STATUS_FILL = {
    'Open':   PatternFill(fill_type='solid', fgColor='FFFFC000'),
    'Closed': PatternFill(fill_type='solid', fgColor='FF92D050'),
}

ISSUES = [
    {
        'date': '2026-07-29',
        'description': 'Meeting Room cable tray routing conflicts with existing HVAC duct',
        'risk': 'Medium',
        'solution': 'Coordinate with MEP to reroute tray, propose alternative by 2026-07-31',
        'action_by': 'Cailleach',
        'status': 'Open',
    },
]

RFAS = [
    {
        'date': '2026-07-30',
        'type': 'RFI',
        'description': 'Request confirmation of final faceplate color (white vs ivory)',
        'issued_to': 'Aurora Tech PM',
        'respond_by': '2026-08-02',
        'status': 'Open',
        'remarks': '',
    },
]

PHOTO_PLACEHOLDERS = [
    {'sub_item': 'Reception',     'phase': 'Cable Pulling', 'description': 'Before cable pulling / 穿线前'},
    {'sub_item': 'Reception',     'phase': 'Cable Pulling', 'description': 'Cable pulling completed / 穿线完成'},
    {'sub_item': 'Reception',     'phase': 'Termination',   'description': 'Termination in progress / 端接过程'},
    {'sub_item': 'Reception',     'phase': 'Termination',   'description': 'Termination completed / 端接完成'},
    {'sub_item': 'Open Office',   'phase': 'Cable Pulling', 'description': 'Before cable pulling / 穿线前'},
    {'sub_item': 'Open Office',   'phase': 'Cable Pulling', 'description': 'Cable pulling completed / 穿线完成'},
    {'sub_item': 'Meeting Room',  'phase': 'Cable Pulling', 'description': 'Before cable pulling / 穿线前'},
]

DATE_NF = '[$-409]d\\-mmm;@'
NO_FILL = PatternFill(fill_type=None)


def parse_date(s):
    if not s:
        return None
    return datetime.strptime(s, '%Y-%m-%d')


def clear_values(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def clear_fill(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.fill = NO_FILL


def update_progress_report(wb):
    ws = wb['Progress Report']
    n_phases = len(PHASES)
    last_phase_col_idx = 2 + n_phases * 3          # 2 + 12 = 14 (N)
    last_phase_col = get_column_letter(last_phase_col_idx)
    first_extra_col = last_phase_col_idx + 1        # 15 (O)
    last_extra_col = 23                              # W

    # --- Metadata (Step 4 of main agent) ---
    ws['E9'] = REPORT_DATE
    ws['E10'] = PROJECT_TITLE

    # --- Unmerge template merges that no longer fit ---
    to_unmerge = []
    for mr in list(ws.merged_cells.ranges):
        s = str(mr)
        # Row 12 title merge (C12:W12)
        if mr.min_row == 12 and mr.max_row == 12:
            to_unmerge.append(s)
        # Row 13 phase header merges for phases beyond our count
        if mr.min_row == 13 and mr.max_row == 13:
            phase_idx = (mr.min_col - 3) // 3
            if phase_idx >= n_phases:
                to_unmerge.append(s)
    for s in to_unmerge:
        ws.unmerge_cells(s)

    # --- Clear extra phase columns (O-W) for rows 12..23 ---
    clear_values(ws, 12, 23, first_extra_col, last_extra_col)
    # Remove leftover blue fill from removed phase headers (row 13) & title bar (row 12)
    clear_fill(ws, 12, 12, first_extra_col, last_extra_col)
    clear_fill(ws, 13, 13, first_extra_col, last_extra_col)

    # --- Row 12: merged title = project.title across C12:{last}12 ---
    ws['C12'] = PROJECT_TITLE
    ws.merge_cells(f'C12:{last_phase_col}12')

    # --- Row 13: phase headers (merge 3 cols each) ---
    for i, phase in enumerate(PHASES):
        col_idx = 3 + i * 3
        col = get_column_letter(col_idx)
        ws[f'{col}13'] = phase
        merge_range = f'{col}13:{get_column_letter(col_idx + 2)}13'
        already = any(str(mr) == merge_range for mr in ws.merged_cells.ranges)
        if not already:
            ws.merge_cells(merge_range)

    # --- Row 14: sub-headers (% / Till Date / Target Date) ---
    for i in range(n_phases):
        base = 3 + i * 3
        ws.cell(row=14, column=base).value = '%'
        ws.cell(row=14, column=base + 1).value = 'Till Date'
        ws.cell(row=14, column=base + 2).value = 'Target Date'

    # --- Row 16: floor identifier ---
    ws['B16'] = PROJECT_FLOOR

    # --- Rows 17..19: sub-item data ---
    phase_cols = [get_column_letter(3 + j * 3) for j in range(n_phases)]
    for i, sub_item in enumerate(SUB_ITEMS):
        row = 17 + i
        ws.cell(row=row, column=1).value = i + 1                  # A: item no.
        ws.cell(row=row, column=2).value = sub_item               # B: description
        for j, phase in enumerate(PHASES):
            base_col = 3 + j * 3
            data = PROGRESS[sub_item][phase]
            pct_cell = ws.cell(row=row, column=base_col)
            pct_cell.value = data['pct'] / 100
            pct_cell.number_format = '0%'
            pct_cell.font = Font(name='Arial', size=10, bold=False,
                                 color=STATUS_FONT_COLOR[data['status']])
            # Till Date — always set (clears leftover template date when empty)
            till_cell = ws.cell(row=row, column=base_col + 1)
            till = parse_date(data['till'])
            if till:
                till_cell.value = till
                till_cell.number_format = DATE_NF
            else:
                till_cell.value = None
            # Target Date — always set
            tgt_cell = ws.cell(row=row, column=base_col + 2)
            target = parse_date(data['target'])
            if target:
                tgt_cell.value = target
                tgt_cell.number_format = DATE_NF
            else:
                tgt_cell.value = None
        # X column: AVERAGE formula across phase % cells
        formula = '=AVERAGE(' + ','.join(f'{c}{row}' for c in phase_cols) + ')'
        ws.cell(row=row, column=24).value = formula               # X = col 24

    # --- Clear leftover 4th sub-item row (row 20) ---
    clear_values(ws, 20, 20, 1, 24)

    # --- Overall row (row 23): clear phase 5-7 cols + clear phase 1-4 dates ---
    clear_values(ws, 23, 23, first_extra_col, last_extra_col)
    for j in range(n_phases):
        base_col = 3 + j * 3
        ws.cell(row=23, column=base_col + 1).value = None        # Till date
        ws.cell(row=23, column=base_col + 2).value = None        # Target date
    # Keep C23/F23/I23/L23 =AVERAGE(C15:C22) etc. (still valid for 3 sub-items)
    # Keep X23 =AVERAGE(X15:X22)

    # --- Clear leftover phase 5-7 averages row fills ---
    clear_fill(ws, 13, 13, first_extra_col, last_extra_col)


def update_site_photo(wb):
    ws = wb['Site Photo']

    # Metadata
    ws['D9'] = REPORT_DATE
    ws['D10'] = PROJECT_TITLE

    # Clear template rows 13-23 (11 placeholder rows)
    clear_values(ws, 13, 23, 1, 12)

    # Fill 7 photo placeholders (rows 13-19)
    report_dt = parse_date(REPORT_DATE)
    for i, ph in enumerate(PHOTO_PLACEHOLDERS):
        row = 13 + i
        ws.cell(row=row, column=1).value = i + 1                  # A: item no.
        b = ws.cell(row=row, column=2)                            # B: date
        b.value = report_dt
        b.number_format = DATE_NF
        ws.cell(row=row, column=3).value = ph['description']      # C: description
        # D: photo placeholder (left blank)


def update_issue_rfa_log(wb):
    ws = wb['Issue_RFA Log']

    # Metadata
    ws['C11'] = PROJECT_TITLE

    # Clear template issue rows 14-17 (4 issues)
    clear_values(ws, 14, 17, 1, 10)

    # Fill issue rows (starting row 14)
    for i, issue in enumerate(ISSUES):
        row = 14 + i
        ws.cell(row=row, column=1).value = i + 1                  # A: item no.
        b = ws.cell(row=row, column=2)                            # B: date
        b.value = parse_date(issue['date'])
        b.number_format = DATE_NF
        ws.cell(row=row, column=3).value = issue['description']   # C: description (merged C:D)
        ws.cell(row=row, column=5).value = issue['risk']          # E: risk
        ws.cell(row=row, column=6).value = issue['solution']      # F: solution
        ws.cell(row=row, column=7).value = issue['action_by']     # G: action by
        i_cell = ws.cell(row=row, column=9)                       # I: status
        i_cell.value = issue['status']
        i_cell.fill = STATUS_FILL.get(issue['status'], NO_FILL)

    # RFI / RFA LOG title (row 18) — kept from template
    # Project row (row 20)
    ws['C20'] = PROJECT_TITLE
    # RFA header (row 22) — kept from template

    # Clear RFA data rows 23+ (template has 1 empty row at 23)
    clear_values(ws, 23, 26, 1, 10)

    # Fill RFA rows (starting row 23)
    for i, rfa in enumerate(RFAS):
        row = 23 + i
        ws.cell(row=row, column=1).value = i + 1                  # A: item no.
        b = ws.cell(row=row, column=2)                            # B: issued date
        b.value = parse_date(rfa['date'])
        b.number_format = DATE_NF
        ws.cell(row=row, column=3).value = rfa['type']           # C: RFI/RFA
        ws.cell(row=row, column=5).value = rfa['description']     # E: description
        ws.cell(row=row, column=6).value = rfa['issued_to']       # F: issued to
        g = ws.cell(row=row, column=7)                            # G: respond by
        g.value = parse_date(rfa['respond_by'])
        g.number_format = DATE_NF
        i_cell = ws.cell(row=row, column=9)                       # I: status
        i_cell.value = rfa['status']
        i_cell.fill = STATUS_FILL.get(rfa['status'], NO_FILL)
        if rfa.get('remarks'):
            ws.cell(row=row, column=10).value = rfa['remarks']    # J: remarks


def main():
    shutil.copy(TEMPLATE, OUTPUT)
    wb = openpyxl.load_workbook(OUTPUT)
    update_progress_report(wb)
    update_site_photo(wb)
    update_issue_rfa_log(wb)
    wb.save(OUTPUT)
    print(f'Saved: {OUTPUT}')


if __name__ == '__main__':
    main()
