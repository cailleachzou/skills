"""Test runner: generate delivery order for Aurora Tech (fictional project)."""
import sys
import os

sys.path.insert(0, r'c:\Users\59620\.claude\skills\tendo-brand\scripts')
from gen_delivery_order import gen_delivery_order
import openpyxl
from datetime import datetime

SKILL_DIR = r'c:\Users\59620\.claude\skills\tendo-brand'
OUTPUT_DIR = os.path.join(SKILL_DIR, 'test')

data = {
    "requisition_no": "TCMR2607-00030",
    "company": "Aurora Tech Shanghai Co., Ltd.",
    "address": "IFC Tower 2, Level 28, Unit 2801, 8 Century Avenue, Pudong New Area, Shanghai, 200120",
    "sales_order_no": "TCSO2607-00200",
    "sales_quotation_no": "TCSQ2607-00255R1",
    "submitted_by": "Cailleach.Zou",
    "date": "2026-07-30",
    "currency": "CNY",
    "items": [
        {"part_no": "760191940", "description": "Faceplate 2-Port, White", "qty": 50, "unit": "个", "unit_cost": 10.00},
        {"part_no": "760192020", "description": "Faceplate 1-Port, White", "qty": 20, "unit": "个", "unit_cost": 8.50},
        {"part_no": "1077354-01", "description": "Cat6 Cable, Blue, 305m/box", "qty": 8, "unit": "箱", "unit_cost": 520.00},
        {"part_no": "2060421-R1", "description": "24-Port Patch Panel, 1U", "qty": 4, "unit": "个", "unit_cost": 380.00},
    ],
}

# Step 1: generate xlsx via shared helper
xlsx_path = gen_delivery_order(data, OUTPUT_DIR)

# Step 1b: update A50 signature date per agent instruction (Field Mapping row 42)
wb = openpyxl.load_workbook(xlsx_path)
ws = wb['Material Requisition']
g5_val = ws['G5'].value
if isinstance(g5_val, datetime):
    ws['A50'] = "Date:" + g5_val.strftime("%d/%m/%y")
else:
    ws['A50'] = "Date:" + str(g5_val)
wb.save(xlsx_path)
print(f"XLSX: {xlsx_path}")

# Step 2: xlsx -> PDF (Excel COM)
pdf_path = None
try:
    import win32com.client
    xl = win32com.client.Dispatch('Excel.Application')
    xl.Visible = False
    xl.DisplayAlerts = False
    xl_wb = xl.Workbooks.Open(os.path.abspath(xlsx_path))
    pdf_path = xlsx_path.rsplit('.', 1)[0] + '.pdf'
    xl_wb.ExportAsFixedFormat(0, pdf_path)  # 0 = xlTypePDF
    xl_wb.Close(False)
    xl.Quit()
    print(f"PDF: {pdf_path}")
except Exception as e:
    print(f"PDF conversion failed: {e}")

# Step 3: overlay signature image (PyMuPDF) — text already in xlsx
if pdf_path:
    try:
        import fitz
        sig_img = os.path.join(SKILL_DIR, 'assets', 'cailleach.png')
        doc = fitz.open(pdf_path)
        page = doc[0]
        sig_instances = page.search_for('Signature:')
        sig_left = sig_instances[1] if len(sig_instances) > 1 else sig_instances[0]
        name_instances = page.search_for('Name:')
        name_left = name_instances[1] if len(name_instances) > 1 else name_instances[0]
        box_top = sig_left.y1 + 1
        box_bottom = name_left.y0 - 2
        box_left = sig_left.x0 - 2
        box_right = 185
        box_center_x = (box_left + box_right) / 2
        img_w = 80
        img_h = int(img_w * 50 / 99)
        img_x = box_center_x - img_w / 2
        img_y = (box_top + box_bottom) / 2 - img_h / 2
        sig_rect = fitz.Rect(img_x, img_y, img_x + img_w, img_y + img_h)
        page.insert_image(sig_rect, filename=sig_img)
        doc.save(pdf_path, incremental=True, encryption=0)
        doc.close()
        print(f"Signature overlaid: {pdf_path}")
    except Exception as e:
        print(f"Signature overlay failed: {e}")

# Step 4: verify xlsx fields
wb = openpyxl.load_workbook(xlsx_path)
ws = wb['Material Requisition']

results = []

# Header fields
header_expected = {
    'G4': data['requisition_no'],
    'C5': data['company'],
    'C6': data['address'],
    'C7': data['sales_order_no'],
    'C8': data['sales_quotation_no'],
    'C9': data['submitted_by'],
    'G6': data['currency'],
}
for cell, exp in header_expected.items():
    actual = ws[cell].value
    results.append((cell, actual == exp, repr(actual), repr(exp)))

# G5 date — must be datetime object = 2026-07-30
g5 = ws['G5'].value
g5_ok = isinstance(g5, datetime) and g5.strftime("%Y-%m-%d") == "2026-07-30"
results.append(('G5', g5_ok, repr(g5), 'datetime 2026-07-30'))

# A50 signature date — updated from G5
a50 = ws['A50'].value
results.append(('A50', a50 == "Date:30/07/26", repr(a50), 'Date:30/07/26'))

# Items (4 rows, 12-15)
for i, item in enumerate(data['items']):
    row = 12 + i
    checks = [
        (f'B{row}', ws.cell(row=row, column=2).value, item['part_no']),
        (f'C{row}', ws.cell(row=row, column=3).value, item['description']),
        (f'E{row}', ws.cell(row=row, column=5).value, item['qty']),
        (f'F{row}', ws.cell(row=row, column=6).value, item['unit']),
        (f'G{row}', ws.cell(row=row, column=7).value, item['unit_cost']),
    ]
    for cell, actual, exp in checks:
        results.append((cell, actual == exp, repr(actual), repr(exp)))

# Row 16 (5th item row) should be empty — not filled
b16_empty = ws['B16'].value in (None, '')
results.append(('B16(empty)', b16_empty, repr(ws['B16'].value), 'None'))

# Formulas preserved (A column sequence + H column totals + G7/G8/G9)
formula_expected = {
    'A12': 1,
    'A13': '=A12+1',
    'A14': '=A13+1',
    'A15': '=A14+1',
    'A16': '=A15+1',
    'H12': '=E12*G12',
    'H13': '=E13*G13',
    'H14': '=E14*G14',
    'H15': '=E15*G15',
    'G7': '=SUM(H12:H37)/1.13',
    'G8': '=G7*0.13',
    'G9': '=SUM(G7:G8)',
}
for cell, exp in formula_expected.items():
    actual = ws[cell].value
    results.append((cell, actual == exp, repr(actual), repr(exp)))

# Signature labels A48/A49 untouched
results.append(('A48', ws['A48'].value == 'Name:Cailleach Zou', repr(ws['A48'].value), 'Name:Cailleach Zou'))
results.append(('A49', ws['A49'].value == 'Title:Senior Project Engineer', repr(ws['A49'].value), 'Title:Senior Project Engineer'))

# Print verification
print("\n=== Verification ===")
all_pass = True
for cell, ok, actual, exp in results:
    status = "PASS" if ok else "FAIL"
    if not ok:
        all_pass = False
    print(f"{status} {cell}: actual={actual} expected={exp}")

pass_n = sum(1 for _, ok, _, _ in results if ok)
fail_n = sum(1 for _, ok, _, _ in results if not ok)
print(f"\nTotal: {len(results)} | Pass: {pass_n} | Fail: {fail_n}")
print(f"Overall: {'PASS' if all_pass else 'FAIL'}")
