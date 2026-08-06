"""Verify generated Aurora Tech Q&A sheet against template structure."""
from pathlib import Path
import openpyxl

BASE = Path(r"c:\Users\59620\.claude\skills\tendo-brand")
TEMPLATE = BASE / "references" / "TendoCN - Q&A for (Client Name) (Project Name) - (Date).xlsx"
OUTPUT = BASE / "test" / "TendoCN - Q&A for Aurora Tech L28 Smart Office Retrofit - 30th Jul'26.xlsx"

EXPECTED_A3 = "Aurora Tech Shanghai L28 Smart Office Retrofit"
EXPECTED_A4 = "Survey Date: 30th Jul'26"

# 模板固定结构位置（来自 agent 指令）
HEADER_ROW8 = ["S/N", "Tendo", "(Client) to Confirm", "Remark"]
CATEGORY_ROWS = [9, 21, 29, 34, 43]      # 合并行，分类标题
SUBTITLE_CELLS = ["B22", "B25", "B30", "B35", "B40"]

results = []

# 读取生成文件
wb = openpyxl.load_workbook(OUTPUT)
ws = wb.active

# 读取模板作对照
wb_t = openpyxl.load_workbook(TEMPLATE)
ws_t = wb_t.active

# 1. A3 项目名称
a3_val = ws['A3'].value
a3_pass = a3_val == EXPECTED_A3
results.append(("A3 项目名称", a3_pass, f"actual={a3_val!r}"))

# 2. A4 日期
a4_val = ws['A4'].value
a4_pass = a4_val == EXPECTED_A4
results.append(("A4 日期", a4_pass, f"actual={a4_val!r}"))

# 3. 表头行 Row 8
hdr_pass = all(ws.cell(row=8, column=c).value == HEADER_ROW8[c-1] for c in range(1, 5))
hdr_detail = [ws.cell(row=8, column=c).value for c in range(1, 5)]
results.append(("表头行 Row8", hdr_pass, f"actual={hdr_detail}"))

# 4. 分类标题行内容应与模板一致（不被破坏）
cat_pass = True
cat_detail = []
for r in CATEGORY_ROWS:
    v_out = ws.cell(row=r, column=1).value
    v_tpl = ws_t.cell(row=r, column=1).value
    if v_out != v_tpl:
        cat_pass = False
    cat_detail.append((r, v_out, v_tpl))
results.append(("分类标题(9/21/29/34/43)", cat_pass, f"out_vs_tpl={cat_detail}"))

# 5. 子标题应与模板一致
sub_pass = True
sub_detail = []
for cell in SUBTITLE_CELLS:
    v_out = ws[cell].value
    v_tpl = ws_t[cell].value
    if v_out != v_tpl:
        sub_pass = False
    sub_detail.append((cell, v_out, v_tpl))
results.append(("子标题(B22/B25/B30/B35/B40)", sub_pass, f"out_vs_tpl={sub_detail}"))

# 6. 合并单元格范围未被破坏
m_out = sorted(str(m) for m in ws.merged_cells.ranges)
m_tpl = sorted(str(m) for m in ws_t.merged_cells.ranges)
merge_pass = m_out == m_tpl
results.append(("合并单元格范围", merge_pass, f"out_count={len(m_out)}, tpl_count={len(m_tpl)}"))
if not merge_pass:
    only_out = set(m_out) - set(m_tpl)
    only_tpl = set(m_tpl) - set(m_out)
    results[-1] = ("合并单元格范围", merge_pass,
                   f"only_in_out={only_out}, only_in_tpl={only_tpl}")

# 7. 数据行（B列）内容应与模板一致 — 抽查几个关键数据行
data_check_rows = [10, 11, 12, 13, 16, 17, 23, 26, 27, 31, 36, 41, 44, 45]
data_pass = True
data_detail = []
for r in data_check_rows:
    v_out = ws.cell(row=r, column=2).value
    v_tpl = ws_t.cell(row=r, column=2).value
    if v_out != v_tpl:
        data_pass = False
        data_detail.append((r, v_out, v_tpl))
results.append(("数据行B列(抽查)", data_pass, f"mismatches={data_detail}"))

print("=" * 60)
print("验证报告")
print("=" * 60)
all_pass = True
for name, ok, detail in results:
    status = "PASS" if ok else "FAIL"
    if not ok:
        all_pass = False
    print(f"[{status}] {name}  -- {detail}")
print("=" * 60)
print(f"总体: {'ALL PASS' if all_pass else 'HAS FAIL'}")
print(f"文件路径: {OUTPUT}")
