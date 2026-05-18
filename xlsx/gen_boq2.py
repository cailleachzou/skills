from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = load_workbook(r'c:\Users\59620\OneDrive\AIproject2\Tendo - 模板文件\Copy of TCSQ2604-00159 Tendo - Provisioning of Servers in HK (6).xlsx')
ws = wb.active

boq_sections = [
    ("1.1 视频安防监控系统（CCTV）", [
        ("网络高清半球摄像机", 34, "台", "400万像素", ""),
        ("网络枪型摄像机", 14, "台", "400万像素", ""),
        ("存储服务器（NVR）", 2, "台", "64路、16盘位", ""),
        ("监控硬盘", 6, "块", "8TB 企业级", ""),
        ("视频管理平台软件", 1, "套", "含门授权", ""),
        ("监视器（操作台）", 1, "套", "55寸 3×3拼接屏", ""),
    ]),
    ("1.2 综合布线系统（SCS）", [
        ("数据信息点（双口面板）", 200, "点", "六类非屏蔽", ""),
        ("数据信息点（单口面板）", 100, "点", "六类非屏蔽", ""),
        ("楼层配线架", 18, "套", "24口配线架", "每层2套"),
        ("楼层光纤配线架", 9, "套", "12口光配架", "每层1套"),
        ("核心光纤配线架", 1, "套", "48口", ""),
        ("跳线（语音/数据）", 150, "条", "Cat.6 2m", ""),
        ("光纤跳线", 40, "条", "LC-SC 多模", ""),
        ("机柜（42U）", 12, "台", "含PDU", ""),
        ("PVC管/线槽", 1, "批", "various", ""),
        ("六类网线", 40, "箱", "UTP Cat.6 低烟无卤", ""),
        ("语音大对数线", 5, "箱", "50对", ""),
    ]),
    ("1.3 无线网络系统（Wi-Fi）", [
        ("Wi-Fi 6 放装AP", 50, "台", "室内型、802.11ax", ""),
        ("无线控制器", 1, "台", "AC控制器", ""),
        ("POE交换机", 2, "台", "48口", ""),
        ("楼层接入交换机", 9, "台", "24口", "每层1台"),
        ("核心交换机", 2, "台", "三层交换机", "HA架构"),
        ("路由器", 2, "台", "双机热备", "HA架构"),
    ]),
    ("1.4 门禁一卡通系统（ACS）", [
        ("人脸识别门禁读卡器", 16, "台", "TCP/IP、支持口罩识别", ""),
        ("四门门禁控制器", 4, "台", "TCP/IP 通讯", ""),
        ("磁力锁", 16, "套", "280kg 拉力", ""),
        ("出门按钮", 16, "个", "86型", ""),
        ("门禁管理软件", 1, "套", "含门授权", ""),
    ]),
    ("1.5 入侵报警系统", [
        ("报警主机", 1, "台", "总线制、≥128防区", ""),
        ("双鉴探测器", 30, "个", "红外+微波", ""),
        ("玻璃破碎探测器", 15, "个", "频谱分析型", ""),
        ("紧急报警按钮", 10, "个", "86型", ""),
        ("声光报警器", 4, "个", "室外型", ""),
        ("报警管理软件", 1, "套", "含地图模块", ""),
    ]),
    ("1.6 公共广播与背景音乐系统", [
        ("广播功放", 3, "台", "1000W/650W", ""),
        ("吸顶音箱", 40, "只", "6寸/30W", ""),
        ("壁挂音箱", 15, "只", "20W", ""),
        ("音量控制器", 10, "个", "3档", ""),
        ("广播话筒", 2, "只", "桌面鹅颈", ""),
        ("广播主机/矩阵", 1, "台", "数码矩阵", ""),
        ("消防广播模块", 8, "个", "与火灾报警联动", ""),
    ]),
    ("1.7 电子巡查系统", [
        ("电子巡查棒（离线式）", 4, "支", "NFC 标签打卡", ""),
        ("NFC 巡查点标签", 20, "个", "防水型", ""),
        ("巡查管理软件", 1, "套", "含App及平台", ""),
        ("通讯座（USB）", 2, "个", "巡查棒数据上传", ""),
    ]),
    ("1.8 周界围栏防护系统", [
        ("脉冲电子围栏主机", 4, "套", "4/6线制", ""),
        ("电子围栏防区模块", 8, "个", "含支架/导线", ""),
        ("高压危险警示牌", 20, "块", "夜间反光式", ""),
        ("避雷及接地装置", 1, "套", "接地电阻≤10Ω", ""),
        ("围栏报警管理软件", 1, "套", "含电子地图", ""),
        ("联动模块", 4, "个", "与监控/照明联动", ""),
    ]),
    ("1.9 道闸系统", [
        ("智能道闸栏杆机", 2, "台", "直流无刷、防折断栅栏杆", ""),
        ("车牌识别摄像机", 2, "台", "200万高清+IR-CUT", ""),
        ("车辆检测器", 2, "个", "地感线圈/微波雷达", ""),
        ("道闸控制箱", 2, "个", "室外防水型含防雷", ""),
        ("收费管理系统", 1, "套", "本地+微信/支付宝", ""),
        ("防砸保护装置", 2, "套", "压力波传感器", ""),
        ("平台联网接口", 1, "套", "城市停车平台预留", ""),
    ]),
    ("1.10 信息导引及发布系统", [
        ("P2.5 LED全彩屏", 1, "套", "≥10㎡、含钢结构", ""),
        ("壁挂信息屏", 7, "台", "21.5寸", ""),
        ("发布管理软件", 1, "套", "", ""),
    ]),
    ("1.11 建筑设备监控系统（BAS）", [
        ("BAS服务器/网关", 1, "台", "数据采集网关", ""),
        ("温湿度传感器", 20, "个", "文物区专用（精度±0.3℃）", ""),
        ("通用输入模块", 10, "个", "模拟量采集", ""),
        ("BAS监控软件", 1, "套", "含组态界面", ""),
        ("管线及辅材", 1, "批", "", ""),
    ]),
    ("1.12 无线对讲系统", [
        ("对讲机", 10, "台", "数字+模拟双段", ""),
        ("中继台", 1, "台", "数字中继", ""),
        ("吸顶天线", 3, "副", "室内全向", ""),
        ("天线馈线及辅材", 1, "批", "", ""),
    ]),
    ("1.13 智能展陈感应解说系统", [
        ("毫米波雷达传感器", 25, "个", "存在检测、博物馆级", "展厅核心展位感应触发"),
        ("多媒体控制主机", 2, "台", "", "冗余热备"),
        ("展柜音箱/嵌入式喇叭", 25, "只", "展柜专用", ""),
        ("互动显示屏", 5, "台", "32寸电容触摸", ""),
        ("多语言语音播控服务器", 1, "套", "中文普通话+英语", "多语言讲解支持"),
        ("4K媒体播放器", 8, "台", "HDMI/网络流媒体输出", "B1×3展厅+1F+2F+3F书画/瓷器/其他展厅"),
        ("地面投影设备", 3, "套", "短焦激光投影、防眩光、防踩踏", "B1起点段+1F段+2F段 时间动线"),
        ("墙面投影设备", 3, "套", "激光投影、含投影幕布", "B1~1F/1F~2F/2F~3F走廊墙面"),
        ("沉浸式LED大屏", 1, "套", "P2.5、含钢结构及控制系统", "3F高潮点"),
        ("声光中控系统", 1, "套", "照明联动+投影同步控制", "时间动线灯光/投影统一控制"),
        ("系统集成及调试", 1, "套", "", ""),
    ]),
    ("1.14 客流统计系统", [
        ("双目客流统计摄像机", 3, "台", "双向统计", ""),
        ("客流分析软件", 1, "套", "", ""),
    ]),
    ("1.15 会议系统", [
        ("音视频预留在2F拍卖厅", 1, "套", "管线+接口面板", ""),
        ("音视频预留3F展厅", 1, "套", "管线+接口面板", ""),
        ("1F接待厅会议系统", 1, "套", "扩声+显示+接口", ""),
    ]),
    ("1.16 机房工程", [
        ("UPS不间断电源", 1, "套", "10KVA/2H", ""),
        ("防静电地板", 40, "㎡", "600×600", ""),
        ("机柜配电单元（PDU）", 9, "个", "16A", ""),
        ("机房动环监控系统", 1, "套", "", ""),
        ("机房装修（简标）", 1, "批", "吊顶/墙面/照明", ""),
    ]),
    ("1.17 综合管理平台", [
        ("综合管理平台软件", 1, "套", "安防集成", ""),
        ("工作站（管理电脑）", 2, "台", "I7/16G/512G", ""),
    ]),
    ("2.1 桥架", [
        ("弱电桥架（主路由）", 200, "米", "200×100", "垂直主干"),
        ("弱电桥架（分支）", 500, "米", "100×50", "各层水平布线"),
        ("桥架配件", 1, "批", "弯头/三通/封口", ""),
    ]),
    ("2.2 管材", [
        ("PVC线管", 2000, "米", "DN25", "楼层水平暗敷段"),
        ("PVC线管", 1000, "米", "DN20", "短距离分支"),
        ("KBG金属管", 600, "米", "DN25", "B1～3F展厅/文物区暗敷段"),
        ("钢管", 200, "米", "DN40", "室外/过墙/楼板穿越段"),
    ]),
    ("2.3 信号线", [
        ("光缆（垂直主干）", 300, "米", "12芯单模", "9层垂直主干，含备用"),
        ("广播喇叭线", 1000, "米", "RVS 2×1.0", "55只音箱"),
        ("报警信号总线", 1500, "米", "RVVP 2×1.0", "入侵报警探测器回路"),
        ("门禁信号线", 1500, "米", "RVVP 4×0.5", "门禁读卡器+出门按钮"),
        ("控制信号线", 2000, "米", "RVVP 2×0.75", "BAS传感器/控制模块"),
    ]),
    ("2.4 电源线", [
        ("电源线", 800, "米", "ZR-RVV 3×2.5", "各系统设备供电"),
        ("UPS配电线", 50, "米", "ZR-RVV 3×10", "机房UPS至各层UPS分配箱"),
    ]),
]

BLUE_FILL  = PatternFill("solid", fgColor="0099FF")
DATA_FONT  = Font(name="Arial", size=10)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# Unmerge ALL merged cells in rows 17 onwards (clearing area + potential conflict zone)
to_unmerge = []
for merge in ws.merged_cells.ranges:
    if merge.min_row >= 17:
        to_unmerge.append(str(merge))
for m in to_unmerge:
    ws.unmerge_cells(m)

# Clear ALL rows 17-144 to remove old data
for r in range(17, 145):
    for c in range(1, 29):
        try:
            ws.cell(row=r, column=c).value = None
        except:
            pass

# Update header info
ws["G7"]  = "DES-2026-JYM-BOQ-001"
ws["D9"]  = "江阴蔡氏博物馆"
ws["D10"] = "江阴蔡氏博物馆 Jiangyin Museum"
ws["G9"]  = "DUDU"
ws["G10"] = "Tendo Technology (Shanghai) Co., Ltd."
ws["G11"] = "+86 183 2126 0125"
ws["G12"] = "2026-04-27"

# Clear old contact info still in D column (B:C merged, actual data in D)
ws.cell(row=11, column=4).value = None  # old telephone (D11)
ws.cell(row=12, column=4).value = None  # old email (D12)

# Update column headers
ws["B14"].value = "No."
ws["C14"].value = "Part No."
ws["D14"].value = "设备名称 / Equipment Name"
ws["E14"].value = "规格 / Specification"
ws["F14"].value = "数量 / Qty"
ws["G14"].value = "单位"
ws["H14"].value = "备注 / Remarks"

# Update RE:
ws["D15"] = "RE: DES-2026-JYM 弱电系统设备清单 / ELV System BOQ"

# Clear old item rows 17-30
for r in range(17, 31):
    for c in range(1, 29):
        ws.cell(row=r, column=c).value = None

# Write BOQ data
current_row = 17
item_counter = 1

for section_name, items in boq_sections:
    # Section header row (blue, merged B-I)
    sec_cell = ws.cell(row=current_row, column=2)
    sec_cell.value = section_name
    sec_cell.font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    sec_cell.fill  = BLUE_FILL
    sec_cell.alignment = LEFT
    ws.row_dimensions[current_row].height = 15
    ws.merge_cells(start_row=current_row, start_column=2,
                   end_row=current_row,   end_column=8)
    current_row += 1

    for name, spec, qty, unit, remark in items:
        # B: No.
        c = ws.cell(row=current_row, column=2)
        c.value = item_counter
        c.font  = DATA_FONT
        c.alignment = CENTER

        # D: equipment name
        c = ws.cell(row=current_row, column=4)
        c.value = name
        c.font  = DATA_FONT
        c.alignment = LEFT

        # E: spec
        c = ws.cell(row=current_row, column=5)
        c.value = spec
        c.font  = DATA_FONT
        c.alignment = LEFT

        # F: qty
        c = ws.cell(row=current_row, column=6)
        c.value = qty
        c.font  = DATA_FONT
        c.alignment = CENTER

        # G: unit
        c = ws.cell(row=current_row, column=7)
        c.value = unit
        c.font  = DATA_FONT
        c.alignment = CENTER

        # H: remarks (单价/合价已移除)
        c = ws.cell(row=current_row, column=8)
        c.value = remark
        c.font  = DATA_FONT
        c.alignment = LEFT

        item_counter += 1
        current_row += 1

# Column widths
ws.column_dimensions["B"].width = 5
ws.column_dimensions["C"].width = 5
ws.column_dimensions["D"].width = 28
ws.column_dimensions["E"].width = 35
ws.column_dimensions["F"].width = 8
ws.column_dimensions["G"].width = 8
ws.column_dimensions["H"].width = 45

out = r"c:\Users\59620\OneDrive\AIproject2\Tendo - 5_项目 Projects\Tendo - DES-2026-JYM 江阴博物馆 Jiangyin Museum\Tendo - 02_预算 Budget & BOQ\DES-2026-JYM-BOQ-001.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Last row used: {current_row - 1}")
print(f"Total items: {item_counter - 1}")
