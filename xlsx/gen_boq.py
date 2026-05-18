from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = load_workbook(r'c:\Users\59620\OneDrive\AIproject2\Tendo - 模板文件\Copy of TCSQ2604-00159 Tendo - Provisioning of Servers in HK (6).xlsx')
ws = wb.active

# Unmerge all merged cells in rows 14 onwards (header/cost area + data area)
to_unmerge = []
for merge in ws.merged_cells.ranges:
    if merge.min_row >= 14:
        to_unmerge.append(str(merge))
for m in to_unmerge:
    ws.unmerge_cells(m)

boq_sections = [
    ("1.1 视频安防监控系统（CCTV）", [
        ("网络高清半球摄像机", "400万像素", "台", 34, ""),
        ("网络枪型摄像机", "400万像素", "台", 14, ""),
        ("存储服务器（NVR）", "64路、16盘位", "台", 2, ""),
        ("监控硬盘", "8TB 企业级", "块", 6, ""),
        ("视频管理平台软件", "含门授权", "套", 1, ""),
        ("监视器（操作台）", "55寸 3×3拼接屏", "套", 1, ""),
    ]),
    ("1.2 综合布线系统（SCS）", [
        ("数据信息点（双口面板）", "六类非屏蔽", "点", 200, ""),
        ("数据信息点（单口面板）", "六类非屏蔽", "点", 100, ""),
        ("楼层配线架", "24口配线架", "套", 18, "每层2套"),
        ("楼层光纤配线架", "12口光配架", "套", 9, "每层1套"),
        ("核心光纤配线架", "48口", "套", 1, ""),
        ("跳线（语音/数据）", "Cat.6 2m", "条", 150, ""),
        ("光纤跳线", "LC-SC 多模", "条", 40, ""),
        ("机柜（42U）", "含PDU", "台", 12, ""),
        ("PVC管/线槽", "various", "批", 1, ""),
        ("六类网线", "UTP Cat.6 低烟无卤", "箱", 40, ""),
        ("语音大对数线", "50对", "箱", 5, ""),
    ]),
    ("1.3 无线网络系统（Wi-Fi）", [
        ("Wi-Fi 6 放装AP", "室内型、802.11ax", "台", 50, ""),
        ("无线控制器", "AC控制器", "台", 1, ""),
        ("POE交换机", "48口", "台", 2, ""),
        ("楼层接入交换机", "24口", "台", 9, "每层1台"),
        ("核心交换机", "三层交换机", "台", 2, "HA架构"),
        ("路由器", "双机热备", "台", 2, "HA架构"),
    ]),
    ("1.4 门禁一卡通系统（ACS）", [
        ("人脸识别门禁读卡器", "TCP/IP、支持口罩识别", "台", 16, ""),
        ("四门门禁控制器", "TCP/IP 通讯", "台", 4, ""),
        ("磁力锁", "280kg 拉力", "套", 16, ""),
        ("出门按钮", "86型", "个", 16, ""),
        ("门禁管理软件", "含门授权", "套", 1, ""),
    ]),
    ("1.5 入侵报警系统", [
        ("报警主机", "总线制、≥128防区", "台", 1, ""),
        ("双鉴探测器", "红外+微波", "个", 30, ""),
        ("玻璃破碎探测器", "频谱分析型", "个", 15, ""),
        ("紧急报警按钮", "86型", "个", 10, ""),
        ("声光报警器", "室外型", "个", 4, ""),
        ("报警管理软件", "含地图模块", "套", 1, ""),
    ]),
    ("1.6 公共广播与背景音乐系统", [
        ("广播功放", "1000W/650W", "台", 3, ""),
        ("吸顶音箱", "6寸/30W", "只", 40, ""),
        ("壁挂音箱", "20W", "只", 15, ""),
        ("音量控制器", "3档", "个", 10, ""),
        ("广播话筒", "桌面鹅颈", "只", 2, ""),
        ("广播主机/矩阵", "数码矩阵", "台", 1, ""),
        ("消防广播模块", "与火灾报警联动", "个", 8, ""),
    ]),
    ("1.7 电子巡查系统", [
        ("电子巡查棒（离线式）", "NFC 标签打卡", "支", 4, ""),
        ("NFC 巡查点标签", "防水型", "个", 20, ""),
        ("巡查管理软件", "含App及平台", "套", 1, ""),
        ("通讯座（USB）", "巡查棒数据上传", "个", 2, ""),
    ]),
    ("1.8 周界围栏防护系统", [
        ("脉冲电子围栏主机", "4/6线制", "套", 4, ""),
        ("电子围栏防区模块", "含支架/导线", "个", 8, ""),
        ("高压危险警示牌", "夜间反光式", "块", 20, ""),
        ("避雷及接地装置", "接地电阻≤10Ω", "套", 1, ""),
        ("围栏报警管理软件", "含电子地图", "套", 1, ""),
        ("联动模块", "与监控/照明联动", "个", 4, ""),
    ]),
    ("1.9 道闸系统", [
        ("智能道闸栏杆机", "直流无刷、防折断栅栏杆", "台", 2, ""),
        ("车牌识别摄像机", "200万高清+IR-CUT", "台", 2, ""),
        ("车辆检测器", "地感线圈/微波雷达", "个", 2, ""),
        ("道闸控制箱", "室外防水型含防雷", "个", 2, ""),
        ("收费管理系统", "本地+微信/支付宝", "套", 1, ""),
        ("防砸保护装置", "压力波传感器", "套", 2, ""),
        ("平台联网接口", "城市停车平台预留", "套", 1, ""),
    ]),
    ("1.10 信息导引及发布系统", [
        ("P2.5 LED全彩屏", "≥10㎡、含钢结构", "套", 1, ""),
        ("壁挂信息屏", "21.5寸", "台", 7, ""),
        ("发布管理软件", "", "套", 1, ""),
    ]),
    ("1.11 建筑设备监控系统（BAS）", [
        ("BAS服务器/网关", "数据采集网关", "台", 1, ""),
        ("温湿度传感器", "文物区专用（精度±0.3℃）", "个", 20, ""),
        ("通用输入模块", "模拟量采集", "个", 10, ""),
        ("BAS监控软件", "含组态界面", "套", 1, ""),
        ("管线及辅材", "", "批", 1, ""),
    ]),
    ("1.12 无线对讲系统", [
        ("对讲机", "数字+模拟双段", "台", 10, ""),
        ("中继台", "数字中继", "台", 1, ""),
        ("吸顶天线", "室内全向", "副", 3, ""),
        ("天线馈线及辅材", "", "批", 1, ""),
    ]),
    ("1.13 智能展陈感应解说系统", [
        ("毫米波雷达传感器", "存在检测、博物馆级", "个", 25, "展厅核心展位感应触发"),
        ("多媒体控制主机", "", "台", 2, "冗余热备"),
        ("展柜音箱/嵌入式喇叭", "展柜专用", "只", 25, ""),
        ("互动显示屏", "32寸电容触摸", "台", 5, ""),
        ("多语言语音播控服务器", "中文普通话+英语", "套", 1, "多语言讲解支持"),
        ("4K媒体播放器", "HDMI/网络流媒体输出", "台", 8, "B1×3展厅+1F+2F+3F书画/瓷器/其他展厅"),
        ("地面投影设备", "短焦激光投影、防眩光、防踩踏", "套", 3, "B1起点段+1F段+2F段 时间动线"),
        ("墙面投影设备", "激光投影、含投影幕布", "套", 3, "B1~1F/1F~2F/2F~3F走廊墙面"),
        ("沉浸式LED大屏", "P2.5、含钢结构及控制系统", "套", 1, "3F高潮点"),
        ("声光中控系统", "照明联动+投影同步控制", "套", 1, "时间动线灯光/投影统一控制"),
        ("系统集成及调试", "", "套", 1, ""),
    ]),
    ("1.14 客流统计系统", [
        ("双目客流统计摄像机", "双向统计", "台", 3, ""),
        ("客流分析软件", "", "套", 1, ""),
    ]),
    ("1.15 会议系统", [
        ("音视频预留在2F拍卖厅", "管线+接口面板", "套", 1, ""),
        ("音视频预留3F展厅", "管线+接口面板", "套", 1, ""),
        ("1F接待厅会议系统", "扩声+显示+接口", "套", 1, ""),
    ]),
    ("1.16 机房工程", [
        ("UPS不间断电源", "10KVA/2H", "套", 1, ""),
        ("防静电地板", "600×600", "㎡", 40, ""),
        ("机柜配电单元（PDU）", "16A", "个", 9, ""),
        ("机房动环监控系统", "", "套", 1, ""),
        ("机房装修（简标）", "吊顶/墙面/照明", "批", 1, ""),
    ]),
    ("1.17 综合管理平台", [
        ("综合管理平台软件", "安防集成", "套", 1, ""),
        ("工作站（管理电脑）", "I7/16G/512G", "台", 2, ""),
    ]),
    ("2.1 桥架", [
        ("弱电桥架（主路由）", "200×100", "米", 200, "垂直主干"),
        ("弱电桥架（分支）", "100×50", "米", 500, "各层水平布线"),
        ("桥架配件", "弯头/三通/封口", "批", 1, ""),
    ]),
    ("2.2 管材", [
        ("PVC线管", "DN25", "米", 2000, "楼层水平暗敷段"),
        ("PVC线管", "DN20", "米", 1000, "短距离分支"),
        ("KBG金属管", "DN25", "米", 600, "B1～3F展厅/文物区暗敷段"),
        ("钢管", "DN40", "米", 200, "室外/过墙/楼板穿越段"),
    ]),
    ("2.3 信号线", [
        ("光缆（垂直主干）", "12芯单模", "米", 300, "9层垂直主干，含备用"),
        ("广播喇叭线", "RVS 2×1.0", "米", 1000, "55只音箱"),
        ("报警信号总线", "RVVP 2×1.0", "米", 1500, "入侵报警探测器回路"),
        ("门禁信号线", "RVVP 4×0.5", "米", 1500, "门禁读卡器+出门按钮"),
        ("控制信号线", "RVVP 2×0.75", "米", 2000, "BAS传感器/控制模块"),
    ]),
    ("2.4 电源线", [
        ("电源线", "ZR-RVV 3×2.5", "米", 800, "各系统设备供电"),
        ("UPS配电线", "ZR-RVV 3×10", "米", 50, "机房UPS至各层UPS分配箱"),
    ]),
]

BLUE_FILL = PatternFill("solid", fgColor="0099FF")
HEADER_FONT = Font(name="Arial", bold=True, size=10)
DATA_FONT  = Font(name="Arial", size=10)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# Clear old item rows 17-34
for r in range(17, 35):
    for c in range(1, 29):
        ws.cell(row=r, column=c).value = None

# Update header info
ws["H7"]  = "DES-2026-JYM-BOQ-001"
ws["D9"]  = "江阴蔡氏博物馆"
ws["D10"] = "江阴蔡氏博物馆 Jiangyin Museum"
ws["H9"]  = "DUDU"
ws["H10"] = "Tendo Technology (Shanghai) Co., Ltd."
ws["H11"] = "+86 183 2126 0125"
ws["H12"] = "2026-04-27"

# Update column headers
ws["B14"].value = "No."
ws["C14"].value = "Part No."
ws["D14"].value = "设备名称 / Equipment Name"
ws["E14"].value = "规格 / Specification"
ws["F14"].value = "单位"
ws["G14"].value = "数量"
ws["H14"].value = "备注 / Remarks"
ws["I14"].value = None
ws["J14"].value = None

# Update RE:
ws["D15"] = "RE: DES-2026-JYM 弱电系统设备清单 / ELV System BOQ"

# Clear cost/markup header rows 14-15, cols L(12)-AB(28)
for r in [14, 15]:
    for c in range(12, 29):
        ws.cell(row=r, column=c).value = None

# Clear Grand Total
ws["I35"].value = None

# Clear all cost/markup columns for item rows
for r in range(17, 35):
    for c in [9, 10, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28]:
        ws.cell(row=r, column=c).value = None

# Write BOQ data
current_row = 17
item_counter = 1

for section_name, items in boq_sections:
    # Section header row (blue, merged B-H)
    # First unmerge any existing merges that might conflict
    sec_cell = ws.cell(row=current_row, column=2)
    sec_cell.value = section_name
    sec_cell.font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    sec_cell.fill  = BLUE_FILL
    sec_cell.alignment = LEFT
    ws.row_dimensions[current_row].height = 15

    # Merge B through H
    ws.merge_cells(start_row=current_row, start_column=2,
                   end_row=current_row,   end_column=8)
    current_row += 1

    for name, spec, unit, qty, remark in items:
        ws.cell(row=current_row, column=2).value = item_counter
        ws.cell(row=current_row, column=2).font  = DATA_FONT
        ws.cell(row=current_row, column=2).alignment = CENTER

        ws.cell(row=current_row, column=4).value = name
        ws.cell(row=current_row, column=4).font  = DATA_FONT
        ws.cell(row=current_row, column=4).alignment = LEFT

        ws.cell(row=current_row, column=5).value = spec
        ws.cell(row=current_row, column=5).font  = DATA_FONT
        ws.cell(row=current_row, column=5).alignment = LEFT

        ws.cell(row=current_row, column=6).value = unit
        ws.cell(row=current_row, column=6).font  = DATA_FONT
        ws.cell(row=current_row, column=6).alignment = CENTER

        ws.cell(row=current_row, column=7).value = qty
        ws.cell(row=current_row, column=7).font  = DATA_FONT
        ws.cell(row=current_row, column=7).alignment = CENTER

        ws.cell(row=current_row, column=8).value = remark
        ws.cell(row=current_row, column=8).font  = DATA_FONT
        ws.cell(row=current_row, column=8).alignment = LEFT

        item_counter += 1
        current_row += 1

# Column widths
ws.column_dimensions["B"].width = 5
ws.column_dimensions["C"].width = 5
ws.column_dimensions["D"].width = 28
ws.column_dimensions["E"].width = 38
ws.column_dimensions["F"].width = 7
ws.column_dimensions["G"].width = 8
ws.column_dimensions["H"].width = 45

out = r"c:\Users\59620\OneDrive\AIproject2\Tendo - 5_项目 Projects\Tendo - DES-2026-JYM 江阴博物馆 Jiangyin Museum\Tendo - 02_预算 Budget & BOQ\DES-2026-JYM-BOQ-001.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Last row used: {current_row - 1}")
print(f"Total items: {item_counter - 1}")
